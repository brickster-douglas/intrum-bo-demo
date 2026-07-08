# Databricks notebook source
# MAGIC %md
# MAGIC # Config-Driven Report Distribution
# MAGIC **Replaces SAP BusinessObjects Publications / Scheduled Reports**
# MAGIC
# MAGIC This Lakeflow Job reads a recipient configuration table, queries per-recipient
# MAGIC filtered data, renders reports (HTML/CSV/Excel), and delivers via email — exactly
# MAGIC like BO's per-recipient scheduling, but governed, auditable, and extensible.
# MAGIC
# MAGIC **Architecture:**
# MAGIC ```
# MAGIC ┌─────────────────────┐     ┌──────────────────┐     ┌─────────────────┐
# MAGIC │ recipient_config    │────▶│  Lakeflow Job     │────▶│  Email (SMTP)   │
# MAGIC │ (Delta config table)│     │  (this notebook)  │     │  + Audit Log    │
# MAGIC └─────────────────────┘     └──────────────────┘     └─────────────────┘
# MAGIC         WHO gets WHAT              RENDER                   DELIVER
# MAGIC         with WHICH filters         per-recipient            + LOG
# MAGIC ```

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 0: Configuration

# COMMAND ----------

# Configure these for your environment
CATALOG = dbutils.widgets.get("catalog") if "catalog" in [w.name for w in dbutils.widgets.getAll()] else "YOUR_CATALOG"
SCHEMA = dbutils.widgets.get("schema") if "schema" in [w.name for w in dbutils.widgets.getAll()] else "bo_demo"
FULL = f"{CATALOG}.{SCHEMA}"

# Email config — store credentials in Databricks Secrets
# Supports Gmail API (OAuth) or SMTP (SendGrid, SES, Office 365)
GMAIL_CLIENT_ID = dbutils.secrets.get(scope="bo-demo", key="gmail-client-id")
GMAIL_CLIENT_SECRET = dbutils.secrets.get(scope="bo-demo", key="gmail-client-secret")
GMAIL_REFRESH_TOKEN = dbutils.secrets.get(scope="bo-demo", key="gmail-refresh-token")
SENDER_EMAIL = "YOUR_EMAIL@example.com"

# SAFETY: Only send to this email during testing. All other recipients get DRY_RUN.
LIVE_RECIPIENT_EMAIL = "YOUR_TEST_EMAIL@example.com"

# Set to False to actually send emails (only to LIVE_RECIPIENT_EMAIL)
DRY_RUN = True

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 1: Verify Recipient Config Table
# MAGIC This is the **control plane** — a governed Delta table that defines WHO gets WHAT report
# MAGIC with WHICH filters. Exactly like BO's publication recipient list, but version-controlled
# MAGIC and auditable via Unity Catalog.

# COMMAND ----------

display(spark.sql(f"""
SELECT recipient_id, recipient_name, email, company, region,
       report_type, filter_column, filter_value, format, is_active
FROM {FULL}.recipient_config
WHERE is_active = TRUE
ORDER BY region
"""))

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 2: Create the Audit Log Table
# MAGIC Every delivery is logged — who received what, when, with which filters.
# MAGIC This is the governance layer that BO never had natively.

# COMMAND ----------

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {FULL}.delivery_audit_log (
    delivery_id STRING NOT NULL COMMENT 'Unique delivery event ID',
    recipient_id INT NOT NULL COMMENT 'FK to recipient_config',
    recipient_name STRING NOT NULL,
    email STRING NOT NULL,
    report_type STRING NOT NULL,
    region_filter STRING COMMENT 'Region filter applied',
    format STRING NOT NULL COMMENT 'Output format: PDF, CSV, Excel',
    row_count INT COMMENT 'Number of data rows in the report',
    file_size_bytes INT COMMENT 'Size of the generated report',
    status STRING NOT NULL COMMENT 'SUCCESS, FAILED, SKIPPED, DRY_RUN',
    error_message STRING COMMENT 'Error details if status=FAILED',
    delivered_at TIMESTAMP NOT NULL COMMENT 'Delivery timestamp',
    execution_duration_ms INT COMMENT 'Time to generate + deliver'
) COMMENT 'Audit trail for all report deliveries. Query this to answer: who received what, when, and with which data filters.'
""")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 3: Report Generators
# MAGIC Each report type has a generator function that queries the data with the recipient's
# MAGIC filters and returns a DataFrame. This is the per-recipient logic that replaces
# MAGIC BO's universe + filter binding.

# COMMAND ----------

import json
import urllib.parse
from datetime import datetime
from pyspark.sql import functions as F

def generate_portfolio_summary(region_filter: str):
    """Generate Portfolio Summary report — outstanding amounts, collection rates by client.
    Equivalent to a BO WebI report filtered by country/region."""
    query = f"""
    SELECT
        region,
        client_name,
        portfolio_type,
        total_cases,
        total_outstanding_eur,
        collected_eur,
        collection_rate,
        avg_days_to_collect,
        month_date,
        ROUND(collected_eur / NULLIF(total_outstanding_eur, 0) * 100, 1) AS recovery_pct
    FROM {FULL}.portfolio_data
    WHERE region = '{region_filter}'
    ORDER BY total_outstanding_eur DESC
    """
    df = spark.sql(query)

    stats = {
        "total_outstanding": df.agg(F.sum("total_outstanding_eur")).collect()[0][0],
        "total_collected": df.agg(F.sum("collected_eur")).collect()[0][0],
        "avg_collection_rate": df.agg(F.avg("collection_rate")).collect()[0][0],
        "total_cases": df.agg(F.sum("total_cases")).collect()[0][0],
        "portfolios": df.count(),
    }
    return df, stats


def generate_collection_report(region_filter: str):
    """Generate Collection Performance report — monthly KPIs.
    Equivalent to a BO scheduled report with country filter."""
    query = f"""
    SELECT
        region,
        month_date,
        new_cases,
        closed_cases,
        total_collected_eur,
        recovery_rate,
        avg_resolution_days,
        top_performer,
        ROUND(closed_cases / NULLIF(new_cases, 0) * 100, 1) AS closure_pct
    FROM {FULL}.collection_report
    WHERE region = '{region_filter}'
    ORDER BY month_date DESC
    """
    df = spark.sql(query)

    stats = {
        "total_collected": df.agg(F.sum("total_collected_eur")).collect()[0][0],
        "avg_recovery_rate": df.agg(F.avg("recovery_rate")).collect()[0][0],
        "cases_closed": df.agg(F.sum("closed_cases")).collect()[0][0],
        "avg_resolution_days": df.agg(F.avg("avg_resolution_days")).collect()[0][0],
    }
    return df, stats


# Registry — maps report_type values to generator functions
REPORT_GENERATORS = {
    "portfolio_summary": generate_portfolio_summary,
    "collection_report": generate_collection_report,
}

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 4: Report Renderers
# MAGIC Convert DataFrames into HTML, CSV, or Excel — ready for email attachment.
# MAGIC This replaces BO's Crystal Reports rendering engine.

# COMMAND ----------

import io
import csv

def render_html_report(df, stats: dict, recipient: dict) -> str:
    """Render a DataFrame as a styled HTML email body with company branding.
    Brand colors: Primary Purple #8750E5, Dark Purple #4F1D8D, Footer #290A4A,
    Light Purple #C9B0EF, Off-White #F5F4F2. Font: Arial/Helvetica."""
    name = recipient["recipient_name"]
    company = recipient["company"]
    report = recipient["report_type"].replace("_", " ").title()
    region = recipient["filter_value"]

    rows = df.collect()
    columns = df.columns

    header_row = "".join(
        f'<th style="padding:10px 14px;text-align:left;background:#4F1D8D;'
        f'color:#ffffff;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;'
        f'font-weight:500">'
        f'{c.replace("_"," ").title()}</th>'
        for c in columns
    )

    data_rows = ""
    for i, row in enumerate(rows[:100]):
        bg = "#F5F4F2" if i % 2 else "#ffffff"
        cells = ""
        for c in columns:
            val = row[c]
            if val is None:
                val = "-"
            elif isinstance(val, float):
                if "rate" in c or "pct" in c:
                    val = f"{val:.1%}" if val < 1 else f"{val:.1f}%"
                else:
                    val = f"{val:,.2f}"
            elif isinstance(val, int):
                val = f"{val:,}"
            cells += f'<td style="padding:8px 14px;border-bottom:1px solid #e8e6e3;font-size:13px;color:#252525">{val}</td>'
        data_rows += f'<tr style="background:{bg}">{cells}</tr>'

    kpi_html = ""
    for k, v in stats.items():
        label = k.replace("_", " ").title()
        if isinstance(v, float):
            if "rate" in k:
                v = f"{v:.1%}" if v < 1 else f"{v:.1f}%"
            else:
                v = f"\u20ac{v:,.0f}"
        elif isinstance(v, int):
            v = f"{v:,}"
        elif v is None:
            v = "-"
        kpi_html += f'''
        <div style="flex:1;min-width:130px;background:#F5F4F2;border-radius:8px;padding:16px;text-align:center;border-top:4px solid #8750E5">
            <div style="font-size:11px;color:#707070;text-transform:uppercase;letter-spacing:0.5px;font-weight:500">{label}</div>
            <div style="font-size:24px;font-weight:700;color:#4F1D8D;margin-top:8px">{v}</div>
        </div>'''

    truncation = (
        f'<p style="color:#707070;font-size:12px;margin-top:8px">'
        f'Showing {min(len(rows), 100)} of {len(rows)} rows. Full dataset attached as {recipient["format"]}.</p>'
        if len(rows) > 100 else ""
    )

    html = f"""<!DOCTYPE html>
<html>
<body style="font-family:Arial,Helvetica,sans-serif;max-width:960px;margin:0 auto;padding:0;background:#ffffff">
    <!-- Header -->
    <div style="background:#290A4A;padding:28px 32px;border-radius:0">
        <table style="width:100%" cellpadding="0" cellspacing="0"><tr>
            <td><span style="color:#ffffff;font-size:26px;font-weight:700;letter-spacing:-0.5px">acme collections</span></td>
            <td style="text-align:right"><span style="color:#C9B0EF;font-size:13px">{datetime.now().strftime('%B %d, %Y')}</span></td>
        </tr></table>
    </div>

    <!-- Report Title Bar -->
    <div style="background:#8750E5;padding:20px 32px">
        <h1 style="color:#ffffff;margin:0;font-size:22px;font-weight:500">{report}</h1>
        <p style="color:#e0d0f5;margin:6px 0 0 0;font-size:14px">
            Prepared for <strong style="color:#fff">{name}</strong> ({company}) &bull; Region: <strong style="color:#fff">{region}</strong>
        </p>
    </div>

    <!-- Content -->
    <div style="padding:24px 32px">
        <div style="display:flex;gap:12px;margin-bottom:28px;flex-wrap:wrap">
            {kpi_html}
        </div>
        <table style="width:100%;border-collapse:collapse;border-radius:6px;overflow:hidden">
            <thead><tr>{header_row}</tr></thead>
            <tbody>{data_rows}</tbody>
        </table>
        {truncation}
    </div>

    <!-- Footer -->
    <div style="background:#290A4A;padding:24px 32px;margin-top:16px">
        <p style="color:#C9B0EF;font-size:12px;margin:0">
            <strong style="color:#ffffff">Report Distribution</strong> &mdash; Powered by Databricks Lakeflow
        </p>
        <p style="color:#8a7aaa;font-size:11px;margin:8px 0 0 0">
            Data source: Unity Catalog &bull; Report ID: {recipient['recipient_id']} &bull;
            Filter: {recipient['filter_column']}={recipient['filter_value']}
        </p>
        <p style="color:#8a7aaa;font-size:11px;margin:4px 0 0 0">
            To change your filters, delivery format, or unsubscribe, contact your report administrator.
        </p>
    </div>
</body>
</html>"""
    return html


def render_csv_report(df) -> str:
    """Render a DataFrame as a CSV string for email attachment."""
    rows = df.collect()
    columns = df.columns
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row[c] for c in columns])
    return output.getvalue()


def render_excel_report(df, recipient: dict) -> bytes:
    """Render a DataFrame as an Excel file (bytes) for email attachment.
    Uses pandas to_excel — reliable on all Databricks runtimes including serverless."""
    import pandas as pd

    pdf = df.toPandas()
    pdf.columns = [c.replace("_", " ").title() for c in pdf.columns]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = recipient["report_type"].replace("_", " ").title()[:31]
        pdf.to_excel(writer, index=False, sheet_name=sheet_name)

        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="1B3139", end_color="1B3139", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
        except Exception:
            pass

    return buffer.getvalue()

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 5: Email Delivery
# MAGIC In production, configure SMTP via Databricks Secrets (SendGrid, SES, or Office 365).
# MAGIC For the demo, we simulate delivery and write sample reports to a volume.

# COMMAND ----------

import base64
import uuid
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import urllib.request


def _get_gmail_access_token():
    """Exchange the refresh token for a fresh Gmail API access token."""
    data = urllib.parse.urlencode({
        "client_id": GMAIL_CLIENT_ID,
        "client_secret": GMAIL_CLIENT_SECRET,
        "refresh_token": GMAIL_REFRESH_TOKEN,
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"})
    resp = urllib.request.urlopen(req, timeout=30)
    return json.loads(resp.read())["access_token"]


def _send_via_gmail_api(raw_message):
    """Send a MIME message via the Gmail API. Accepts str or bytes."""
    token = _get_gmail_access_token()
    if isinstance(raw_message, str):
        raw_message = raw_message.encode("utf-8")
    encoded = base64.urlsafe_b64encode(raw_message).decode("ascii")
    payload = json.dumps({"raw": encoded}).encode()
    req = urllib.request.Request(
        "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
        data=payload,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }
    )
    resp = urllib.request.urlopen(req, timeout=30)
    return json.loads(resp.read())


def send_report_email(recipient: dict, html_body: str = None, csv_body: str = None,
                       excel_body: bytes = None, dry_run: bool = True) -> dict:
    """Send (or simulate) an email with the rendered report via Gmail API.
    SAFETY: Only actually sends to LIVE_RECIPIENT_EMAIL. All others get DRY_RUN."""
    start = datetime.now()
    email_to = recipient["email"]
    report_name = recipient["report_type"].replace("_", " ").title()
    subject = f"{report_name} — {recipient['filter_value']} — {datetime.now().strftime('%B %Y')}"

    effective_dry_run = dry_run or (email_to != LIVE_RECIPIENT_EMAIL)

    if effective_dry_run:
        reason = "DRY_RUN mode" if dry_run else f"recipient {email_to} not in live list"
        attachments = []
        if csv_body:
            attachments.append(f"CSV ({len(csv_body):,} bytes)")
        if excel_body:
            attachments.append(f"Excel ({len(excel_body):,} bytes)")

        print(f"    DRY RUN ({reason}): Would send to {email_to}")
        print(f"       Subject: {subject}")
        print(f"       Body: HTML ({len(html_body):,} bytes)" if html_body else "       Body: (none)")
        if attachments:
            print(f"       Attachments: {', '.join(attachments)}")

        result = {
            "status": "DRY_RUN",
            "error_message": None,
            "file_size_bytes": (len(html_body) if html_body else 0) +
                              (len(csv_body) if csv_body else 0) +
                              (len(excel_body) if excel_body else 0),
        }
    else:
        try:
            msg = MIMEMultipart("mixed")
            msg["From"] = SENDER_EMAIL
            msg["To"] = email_to
            msg["Subject"] = subject

            if html_body:
                msg.attach(MIMEText(html_body, "html"))

            if csv_body:
                att = MIMEBase("application", "octet-stream")
                att.set_payload(csv_body.encode("utf-8"))
                encoders.encode_base64(att)
                fname = f"{recipient['report_type']}_{recipient['filter_value']}_{datetime.now().strftime('%Y%m%d')}.csv"
                att.add_header("Content-Disposition", f"attachment; filename={fname}")
                msg.attach(att)

            if excel_body:
                att = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                att.set_payload(excel_body)
                encoders.encode_base64(att)
                fname = f"{recipient['report_type']}_{recipient['filter_value']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                att.add_header("Content-Disposition", f"attachment; filename={fname}")
                msg.attach(att)

            api_result = _send_via_gmail_api(msg.as_bytes())
            print(f"    SENT to {email_to} (Gmail message ID: {api_result.get('id', 'unknown')})")

            result = {"status": "SUCCESS", "error_message": None,
                      "file_size_bytes": len(html_body or "") + len(csv_body or "") + len(excel_body or b"")}
        except Exception as e:
            print(f"    FAILED: {str(e)[:200]}")
            result = {"status": "FAILED", "error_message": str(e)[:500], "file_size_bytes": 0}

    elapsed = int((datetime.now() - start).total_seconds() * 1000)
    result["duration_ms"] = elapsed
    return result

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 6: Main Orchestrator
# MAGIC This is the core loop — iterate over active recipients, generate their personalized
# MAGIC report, deliver it, and log the result. Runs as a scheduled Lakeflow Job.

# COMMAND ----------

def run_report_distribution():
    """Main entry point — called by the Lakeflow Job scheduler."""

    print("=" * 70)
    print(f"  CONFIG-DRIVEN REPORT DISTRIBUTION")
    print(f"  Replacing: SAP BusinessObjects Publications")
    print(f"  Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Mode: {'DRY RUN (demo)' if DRY_RUN else 'LIVE'}")
    print(f"  Data source: {FULL}")
    print("=" * 70)

    recipients_df = spark.sql(f"""
        SELECT * FROM {FULL}.recipient_config
        WHERE is_active = TRUE
        ORDER BY region, recipient_name
    """)
    recipients = [row.asDict() for row in recipients_df.collect()]
    print(f"\n  Active recipients: {len(recipients)}")

    delivery_log = []

    for i, r in enumerate(recipients, 1):
        rid = r["recipient_id"]
        name = r["recipient_name"]
        region = r["filter_value"]
        fmt = r["format"]
        report_type = r["report_type"]

        print(f"\n{'─' * 60}")
        print(f"  [{i}/{len(recipients)}] {name} <{r['email']}>")
        print(f"  Company: {r['company']} | Region: {region}")
        print(f"  Report: {report_type} | Format: {fmt}")

        try:
            generator = REPORT_GENERATORS.get(report_type)
            if not generator:
                raise ValueError(f"Unknown report type: '{report_type}'. "
                               f"Available: {list(REPORT_GENERATORS.keys())}")

            df, stats = generator(region)
            row_count = df.count()
            print(f"  >> Generated: {row_count} rows")

            html_body = render_html_report(df, stats, r)
            csv_body = None
            excel_body = None

            if fmt == "CSV":
                csv_body = render_csv_report(df)
                print(f"  >> Rendered: HTML ({len(html_body):,}B) + CSV ({len(csv_body):,}B)")
            elif fmt == "Excel":
                excel_body = render_excel_report(df, r)
                print(f"  >> Rendered: HTML ({len(html_body):,}B) + Excel ({len(excel_body):,}B)")
            else:
                print(f"  >> Rendered: HTML ({len(html_body):,}B) [PDF would use weasyprint in prod]")

            result = send_report_email(r, html_body, csv_body, excel_body, dry_run=DRY_RUN)

            delivery_log.append({
                "delivery_id": f"DEL-{datetime.now().strftime('%Y%m%d%H%M%S')}-{rid}",
                "recipient_id": rid,
                "recipient_name": name,
                "email": r["email"],
                "report_type": report_type,
                "region_filter": region,
                "format": fmt,
                "row_count": row_count,
                "file_size_bytes": result.get("file_size_bytes", 0),
                "status": result["status"],
                "error_message": result.get("error_message"),
                "delivered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "execution_duration_ms": result["duration_ms"],
            })

        except Exception as e:
            print(f"  >> FAILED: {str(e)[:200]}")
            delivery_log.append({
                "delivery_id": f"DEL-{datetime.now().strftime('%Y%m%d%H%M%S')}-{rid}",
                "recipient_id": rid,
                "recipient_name": name,
                "email": r["email"],
                "report_type": report_type,
                "region_filter": region,
                "format": fmt,
                "row_count": 0,
                "file_size_bytes": 0,
                "status": "FAILED",
                "error_message": str(e)[:500],
                "delivered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "execution_duration_ms": 0,
            })

    if delivery_log:
        from pyspark.sql.types import StructType, StructField, StringType, IntegerType, TimestampType
        schema = StructType([
            StructField("delivery_id", StringType(), False),
            StructField("recipient_id", IntegerType(), False),
            StructField("recipient_name", StringType(), False),
            StructField("email", StringType(), False),
            StructField("report_type", StringType(), False),
            StructField("region_filter", StringType(), True),
            StructField("format", StringType(), False),
            StructField("row_count", IntegerType(), True),
            StructField("file_size_bytes", IntegerType(), True),
            StructField("status", StringType(), False),
            StructField("error_message", StringType(), True),
            StructField("delivered_at", StringType(), False),
            StructField("execution_duration_ms", IntegerType(), True),
        ])
        log_df = spark.createDataFrame(delivery_log, schema=schema)
        log_df = log_df.withColumn("delivered_at", F.to_timestamp("delivered_at"))
        log_df.write.mode("append").saveAsTable(f"{FULL}.delivery_audit_log")
        print(f"\n  Audit log: {len(delivery_log)} entries written to {FULL}.delivery_audit_log")

    success = sum(1 for d in delivery_log if d["status"] in ("SUCCESS", "DRY_RUN"))
    failed = sum(1 for d in delivery_log if d["status"] == "FAILED")

    print(f"\n{'=' * 70}")
    print(f"  DISTRIBUTION COMPLETE")
    print(f"  Delivered: {success} | Failed: {failed} | Total: {len(delivery_log)}")
    print(f"  Run ended: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 70}")

    return delivery_log

# COMMAND ----------

# MAGIC %md
# MAGIC ## Run It!
# MAGIC Execute the full distribution pipeline. In DRY_RUN mode, emails are simulated
# MAGIC but the full pipeline runs end-to-end: config read → query → render → log.

# COMMAND ----------

results = run_report_distribution()

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 7: Verify Audit Log
# MAGIC Every delivery is tracked — answering "who received what, when, with which data?"
# MAGIC This is full compliance-grade audit that BO never had natively.

# COMMAND ----------

display(spark.sql(f"""
SELECT
    delivered_at,
    recipient_name,
    email,
    report_type,
    region_filter,
    format,
    row_count,
    file_size_bytes,
    status,
    execution_duration_ms
FROM {FULL}.delivery_audit_log
ORDER BY delivered_at DESC
LIMIT 20
"""))

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 8: Sample Report Preview
# MAGIC Let's preview what one recipient would see in their email.

# COMMAND ----------

sample = spark.sql(f"SELECT * FROM {FULL}.recipient_config WHERE is_active = TRUE LIMIT 1").collect()[0].asDict()
generator = REPORT_GENERATORS[sample["report_type"]]
df, stats = generator(sample["filter_value"])
html = render_html_report(df, stats, sample)

print(f"Preview for: {sample['recipient_name']} ({sample['email']})")
print(f"Report: {sample['report_type']} | Region: {sample['filter_value']} | Format: {sample['format']}")
print(f"Rows: {df.count()} | HTML size: {len(html):,} bytes")

displayHTML(html)

# COMMAND ----------

# MAGIC %md
# MAGIC ## What This Replaces in SAP BusinessObjects
# MAGIC
# MAGIC | BO Concept | Databricks Equivalent | Advantage |
# MAGIC |---|---|---|
# MAGIC | **Publication** | This Lakeflow Job | Governed, version-controlled, extensible |
# MAGIC | **Recipient list** | `recipient_config` Delta table | Full audit history, RBAC, lineage |
# MAGIC | **Universe filters -> per-user** | SQL WHERE from config table | Dynamic, any column, any logic |
# MAGIC | **Crystal Reports rendering** | HTML/CSV/Excel Python functions | Any format, custom branding |
# MAGIC | **BO scheduling (CMS)** | Databricks Jobs cron scheduler | Serverless, scales to zero |
# MAGIC | **No audit trail** | `delivery_audit_log` Delta table | Compliance-grade: who saw what, when |
# MAGIC
# MAGIC ### Advantages over BO Publications
# MAGIC 1. **Governed config** — recipient table in Unity Catalog with full audit, lineage, RBAC
# MAGIC 2. **Any format** — HTML, CSV, Excel, PDF (with weasyprint). Even Slack/Teams webhooks
# MAGIC 3. **Any data source** — queries run against governed Delta tables, not siloed BO universes
# MAGIC 4. **External recipients** — send to any email, no Databricks/Entra ID identity needed
# MAGIC 5. **Extensible** — add new report types by writing a Python function, not BO Designer
# MAGIC 6. **Auditable** — every delivery logged; compliance can query who received what data
# MAGIC 7. **Cost-efficient** — runs on serverless SQL, scales to zero when idle
# MAGIC
# MAGIC ### To Go Live
# MAGIC 1. Configure SMTP credentials in Databricks Secrets (`dbutils.secrets`)
# MAGIC 2. Set `DRY_RUN = False`
# MAGIC 3. Schedule as a Lakeflow Job (e.g., daily at 07:00, or weekly on Mondays)
# MAGIC 4. Add PDF rendering with `weasyprint` if needed
# MAGIC 5. Monitor via the `delivery_audit_log` table
# MAGIC
# MAGIC ### Migration Path
# MAGIC - **Today:** Use this Lakeflow Job for per-recipient distribution
# MAGIC - **When available:** When Databricks ships native report bursting in AI/BI subscriptions,
# MAGIC   evaluate switching — no data changes needed, just a different distribution method
